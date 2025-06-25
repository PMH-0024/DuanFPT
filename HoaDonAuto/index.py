# import
import os
import time
import shutil
import xml.etree.ElementTree as ET
import pandas as pd
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook, Workbook

#Bước 1: Cấu hình thư mục và trình duyệt
def open_browser(thu_muc_tai_hoa_don):
    # Cấu hình thư mucthu_muc_tai_hoa_don
    os.makedirs(thu_muc_tai_hoa_don, exist_ok=True)
    options = Options()
    options.add_experimental_option("prefs", {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True, 
        "download.default_directory": thu_muc_tai_hoa_don,
        "plugins.always_open_pdf_externally": True,
        "profile.default_content_settings.popups": 0,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1
    })
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)
    return driver, WebDriverWait(driver, 10)
# Bước 2: Tra cứu hóa đơn
def tra_cuu_hoa_don(driver, wait,ma_so_thue, ma_tra_cuu, url ):
    try:    
        driver.get(url)
        # FPT
        if "https://tracuuhoadon.fpt.com.vn/search.html" in url:
            ma_so_thue = str(ma_so_thue).strip().replace("'", "")
            ma_tra_cuu = str(ma_tra_cuu).strip()
            # Nhập mã số thuế
            mst_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='MST bên bán']")))
            driver.execute_script("arguments[0].scrollIntoView(true);", mst_input)
            mst_input.clear()
            mst_input.send_keys(ma_so_thue)
            # Mã tra cứu
            mtc_input = driver.find_element(By.XPATH, "//input[@placeholder='Mã tra cứu hóa đơn']")
            driver.execute_script("arguments[0].scrollIntoView(true);", mtc_input)
            mtc_input.clear()
            mtc_input.send_keys(ma_tra_cuu)
            # Button tra cứu
            print("Đang nhấn nút tra cứu...")
            btn_fpt_search = driver.find_element(By.XPATH,"//button[contains(@class, 'webix_button') and contains(text(), 'Tra cứu')]")
            time.sleep(0.5)
            btn_fpt_search.click()
        # Misa
        elif "https://www.meinvoice.vn/tra-cuu/" in url:
            # Nhập mã hóa đơn
            mtc_input_misa = wait.until(EC.presence_of_element_located((By.NAME, "txtCode")))
            # Đề phòng bị ghi đè
            driver.execute_script("""
                   const header = document.querySelector('.top-header');
                   if (header) header.style.display = 'none';
                   arguments[0].scrollIntoView({block: 'center'});
               """, mtc_input_misa)
            time.sleep(0.5)
            driver.execute_script("arguments[0].value = '';", mtc_input_misa)
            mtc_input_misa.clear()
            mtc_input_misa.send_keys(ma_tra_cuu)
            # Button tra cứu
            print("Đang nhấn nút tra cứu...")
            btn_search_misa = wait.until(EC.element_to_be_clickable((By.ID, "btnSearchInvoice")))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn_search_misa)
            time.sleep(0.5)
            btn_search_misa.click()
        # EHoaDon
        elif "https://van.ehoadon.vn/TCHD?MTC=" in url:
            # Nhập mã hóa đơn
            mts_ehoadon = wait.until(EC.presence_of_element_located((By.ID, "txtInvoiceCode")))
            driver.execute_script("arguments[0].scrollIntoView(true);", mts_ehoadon)
            mts_ehoadon.clear()
            mts_ehoadon.send_keys(ma_tra_cuu)
            # Button tra cứu
            print("Đang nhấn nút tra cứu...")
            btn_ehoadon_search = driver.find_element(By.CLASS_NAME, "btnSearch")
            btn_ehoadon_search.click()
        print("Đang chờ kết quả hóa đơn hiển thị")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
    except TimeoutException():
        print("Tra cứu thất bại")
# Bước 3: Tải hóa đơn XML
def tai_file_xml(driver,wait, thu_muc_tai_hoa_don, url, ma_tra_cuu):
    try:
        # FPT
        if "https://tracuuhoadon.fpt.com.vn/search.html" in url:
            button_fpt = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[span[contains(@class, 'mdi-xml')] and contains(text(), 'Tải XML')]"))
            )
            button_fpt.click()
            print("Đã nhấn nút tải file XML từ FPT")
            time.sleep(2)
        # Misa
        elif "https://www.meinvoice.vn/tra-cuu/" in url:
            btn_misa = wait.until(
                EC.element_to_be_clickable((By.CLASS_NAME, "download"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", btn_misa)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", btn_misa)
            print("Đã nhấn mở menu")
            btn_down_xml_meinvoice = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "txt-download-xml")))
            driver.execute_script("arguments[0].click();", btn_down_xml_meinvoice)
            print("Đã chọn tải hóa đơn file XML")
            time.sleep(2)
        # EHoaDon
        elif "https://van.ehoadon.vn/TCHD?MTC=" in url:
            print("Di chuyển vào iframe của EHoaDon")
            try:
                wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "frameViewInvoice")))
                print("Đã vào iframe")
            except:
                print("Không thể vào iframe")
                return None
            # Bước 1: Di chuyển chuột vào nút download
            btn_down_ehoadon = wait.until( EC.presence_of_element_located((By.ID, "btnDownload")))
            ActionChains(driver).move_to_element(btn_down_ehoadon).perform()
            print("Đã hover nút Download")
            # B2: Hiển thị menu bằng JavaScript
            driver.execute_script("document.querySelector('#divDownloads .dropdown-menu').style.display='block';")
            time.sleep(1)
            # B3: Click vào "Tải XML"
            btn_down_xml_ehoadon = wait.until(EC.element_to_be_clickable((By.ID, "LinkDownXML")))
            btn_down_xml_ehoadon.click()
            print("Đã chọn tải hóa đơn dạng XML")
            time.sleep(2)
            # B4: Thoát khỏi iframe về lại trang chính
            driver.switch_to.default_content()
    except TimeoutException:
        print("Không tìm thấy nút tải XML hoặc mã tra cứu không đúng.")
        return None
    except Exception as e:
        print("Lỗi khi click nút tải:", e)
        return None
        # Tạo thư mục riêng cho từng file hóa đơn
    folder = urlparse(url).netloc.replace("www.", "")
    domain_folder = os.path.join(thu_muc_tai_hoa_don, folder)
    os.makedirs(domain_folder, exist_ok=True)
    # Tìm file XML vừa tải
    for _ in range(10):
        files = os.listdir(thu_muc_tai_hoa_don)
        for file in files:
            if file.endswith(".xml"):
                src = os.path.join(thu_muc_tai_hoa_don, file)
                dest = os.path.join(domain_folder, f"{ma_tra_cuu}.xml")
                shutil.move(src, dest)
                print(f"Đã lưu file XML: {dest}")
                return dest
        time.sleep(1)
    print("Không tìm thấy file XMl vừa tải")
    return None
# Đọc dữ liệu từ file XML
def read_invoice_xml(xml_file_path):
    try:
        # B1: Phân tích XML
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        # B2: Ưu tiên tìm node HDon/DLHDon
        hdon_node = root.find(".//HDon")
        invoice_node = hdon_node.find("DLHDon") if hdon_node is not None else None
        # B3: Nếu không tìm thấy, thử lần lượt các node phổ biến khác
        if invoice_node is None:
            for tag in [".//DLHDon", ".//TDiep", ".//Invoice"]:
                node = root.find(tag)
                if node is not None:
                    invoice_node = node
                    break
            else:
                print(f"Không xác định được node dữ liệu chính trong file: {os.path.basename(xml_file_path)}")
                return None
        # B4: Hàm tìm nhanh theo path kiểu "NDHDon/NBan/Ten"
        def find(path):
            current = invoice_node
            for part in path.split("/"):
                if current is not None:
                    current = current.find(part)
                else:
                    return None
            return current.text if current is not None else None
        # B5: Lấy thông tin số tài khoản bán (có thể nằm trong TTKhac)
        stk_ban = find("NDHDon/NBan/STKNHang")
        if not stk_ban:
            for thongtin in invoice_node.findall(".//NBan/TTKhac/TTin"):
                if thongtin.findtext("TTruong") == "SellerBankAccount":
                    stk_ban = thongtin.findtext("DLieu")
                    break
        # B6: Trả về thông tin cần thiết từ XML
        return {
            'Số hóa đơn': find("TTChung/SHDon"),
            'Đơn vị bán hàng': find("NDHDon/NBan/Ten"),
            'Mã số thuế bán': find("NDHDon/NBan/MST"),
            'Địa chỉ bán': find("NDHDon/NBan/DChi"),
            'Số tài khoản bán': stk_ban,
            'Họ tên người mua hàng': find("NDHDon/NMua/Ten"),
            'Địa chỉ mua': find("NDHDon/NMua/DChi"),
            'Mã số thuế mua': find("NDHDon/NMua/MST"),
        }
    except Exception as error:
        print(f"Lỗi khi đọc file XML {os.path.basename(xml_file_path)}: {error}")
        return None
# Ghi ra thông tin đã đọc vào excel
def append_to_excel(filepath, row_data):
    if not os.path.isfile(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append([
            "STT", "Mã số thuế", "Mã tra cứu", "URL",
            "Số hóa đơn", "Đơn vị bán hàng", "Mã số thuế bán", "Địa chỉ bán", "Số tài khoản bán",
            "Họ tên người mua hàng", "Địa chỉ mua", "Mã số thuế mua"
        ])
        wb.save(filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    ws.append(row_data)
    wb.save(filepath)
# Hàm main chính
def main():
    input_file = "input.xlsx"
    output_file = "output.xlsx"
    thu_muc_tai_hoa_don = os.path.join(os.getcwd(),"InvoiceData")
    driver, wait = open_browser(thu_muc_tai_hoa_don)
    df_invoice = pd.read_excel(input_file, dtype=str)
    for index, row in df_invoice.iterrows():
        stt = index + 1
        ma_so_thue = str(row.get("Mã số thuế", "") or "").strip()
        ma_tra_cuu = str(row.get("Mã tra cứu", "") or "").strip()
        url = str(row.get("URL", "") or "").strip()
        if not url or not ma_tra_cuu:
            continue
        print(f"\n Đang tra cứu mã: {ma_tra_cuu} | Trang web: {url}")
        tra_cuu_hoa_don(driver, wait,ma_so_thue, ma_tra_cuu, url)
        xml_path = tai_file_xml(driver,wait, thu_muc_tai_hoa_don, url, ma_tra_cuu)
        if xml_path:
            parsed = read_invoice_xml(xml_path)
            if parsed:
                row_data = [stt, ma_so_thue, ma_tra_cuu, url] + list(parsed.values()) + [""]
            else:
                row_data = [stt, ma_so_thue, ma_tra_cuu, url] + [""] * 9 + [os.path.basename(xml_path)]
        else:
            row_data = [stt, ma_so_thue, ma_tra_cuu, url] + [""] * 9 + [""]
        append_to_excel(output_file, row_data)
    driver.quit()
    print(f"Đã đọc xong và  lưu lại ở {output_file}")
if __name__ == "__main__":
    main()




